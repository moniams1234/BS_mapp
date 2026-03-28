"""
export_utils.py
Builds Excel workbook and JSON export from analyzed data.
"""
from __future__ import annotations

import io
import json
from datetime import datetime
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Font, PatternFill, Border, Side, numbers
)
from openpyxl.utils import get_column_letter


# ─────────────────────────── helpers ────────────────────────────

def _header_style(ws, row: int, cols: int,
                  bg: str = "1E3A5F", fg: str = "FFFFFF"):
    fill = PatternFill("solid", start_color=bg, end_color=bg)
    font = Font(bold=True, color=fg, name="Arial", size=10)
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _auto_width(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 60)


def _write_df(ws, df: pd.DataFrame, start_row: int = 1):
    """Write a DataFrame to a worksheet starting at start_row."""
    for ci, col_name in enumerate(df.columns, 1):
        ws.cell(row=start_row, column=ci, value=str(col_name))
    _header_style(ws, start_row, len(df.columns))
    for ri, row in enumerate(df.itertuples(index=False), start_row + 1):
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci)
            if isinstance(val, float) and not np.isnan(val):
                cell.value = round(val, 2)
                cell.number_format = '#,##0.00'
            elif isinstance(val, (int, np.integer)):
                cell.value = int(val)
            else:
                cell.value = val if val is not None and str(val) != "nan" else ""


# ─────────────────────────── Excel export ────────────────────────

def build_excel_export(
    raw_df: pd.DataFrame,
    mapp_df: pd.DataFrame,
    bs: dict,
    flags: list[dict],
    filename: str = "",
) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    # ── Raw_Trial_Balance ──────────────────────────────────────────
    ws_raw = wb.create_sheet("Raw_Trial_Balance")
    _write_df(ws_raw, raw_df)
    _auto_width(ws_raw)

    # ── Mapp ──────────────────────────────────────────────────────
    if mapp_df is not None and not mapp_df.empty:
        ws_mapp = wb.create_sheet("Mapp")
        display_cols = ["side", "account_number", "account_name", "group",
                        "debit", "credit", "saldo_dt", "saldo_ct",
                        "persaldo", "persaldo_group", "mapping_status"]
        cols = [c for c in display_cols if c in mapp_df.columns]
        _write_df(ws_mapp, mapp_df[cols])
        _auto_width(ws_mapp)

    # ── Mapping ────────────────────────────────────────────────────
    # Re-write Mapp as "Mapping" (same data, alias)
    if mapp_df is not None and not mapp_df.empty:
        ws_mapping = wb.create_sheet("Mapping")
        _write_df(ws_mapping, mapp_df[cols])
        _auto_width(ws_mapping)

    # ── Balance_Sheet ─────────────────────────────────────────────
    ws_bs = wb.create_sheet("Balance_Sheet")
    row = 1
    if bs:
        # Assets section
        ws_bs.cell(row=row, column=1, value="ASSETS")
        ws_bs.cell(row=row, column=1).font = Font(bold=True, size=12, name="Arial")
        row += 1
        if "assets_detail" in bs and not bs["assets_detail"].empty:
            _write_df(ws_bs, bs["assets_detail"], start_row=row)
            row += len(bs["assets_detail"]) + 2

        ws_bs.cell(row=row, column=1, value="Total Assets")
        ws_bs.cell(row=row, column=2, value=round(bs.get("total_assets", 0), 2))
        ws_bs.cell(row=row, column=1).font = Font(bold=True, name="Arial")
        row += 2

        # Liabilities
        ws_bs.cell(row=row, column=1, value="LIABILITIES")
        ws_bs.cell(row=row, column=1).font = Font(bold=True, size=12, name="Arial")
        row += 1
        if "liabilities_detail" in bs and not bs["liabilities_detail"].empty:
            _write_df(ws_bs, bs["liabilities_detail"], start_row=row)
            row += len(bs["liabilities_detail"]) + 2

        ws_bs.cell(row=row, column=1, value="Total Liabilities")
        ws_bs.cell(row=row, column=2, value=round(bs.get("total_liabilities", 0), 2))
        ws_bs.cell(row=row, column=1).font = Font(bold=True, name="Arial")
        row += 2

        # Equity
        ws_bs.cell(row=row, column=1, value="EQUITY")
        ws_bs.cell(row=row, column=1).font = Font(bold=True, size=12, name="Arial")
        row += 1
        if "equity_detail" in bs and not bs["equity_detail"].empty:
            _write_df(ws_bs, bs["equity_detail"], start_row=row)
            row += len(bs["equity_detail"]) + 2

        ws_bs.cell(row=row, column=1, value="Total Equity")
        ws_bs.cell(row=row, column=2, value=round(bs.get("total_equity", 0), 2))
        ws_bs.cell(row=row, column=1).font = Font(bold=True, name="Arial")
        row += 2

        # Difference
        ws_bs.cell(row=row, column=1, value="BALANCE DIFFERENCE")
        ws_bs.cell(row=row, column=2, value=round(bs.get("difference", 0), 2))
        diff_font = Font(bold=True, name="Arial",
                         color="FF0000" if abs(bs.get("difference", 0)) > 1 else "00AA00")
        ws_bs.cell(row=row, column=1).font = diff_font
        ws_bs.cell(row=row, column=2).font = diff_font

    _auto_width(ws_bs)

    # ── Red_Flags ─────────────────────────────────────────────────
    ws_rf = wb.create_sheet("Red_Flags")
    flags_df = pd.DataFrame(flags or [])
    if not flags_df.empty:
        _write_df(ws_rf, flags_df)
        _auto_width(ws_rf)
    else:
        ws_rf.cell(row=1, column=1, value="No flags generated.")

    # ── Summary ───────────────────────────────────────────────────
    ws_sum = wb.create_sheet("Summary")
    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    summary_data = [
        ("Generated At", now),
        ("Source File", filename or "N/A"),
        ("Total Accounts", len(raw_df) if raw_df is not None else 0),
        ("Mapped Accounts", len(mapp_df[mapp_df["mapping_status"] == "mapped"]) if mapp_df is not None and not mapp_df.empty else 0),
        ("Heuristic Mapped", len(mapp_df[mapp_df["mapping_status"] == "heuristic"]) if mapp_df is not None and not mapp_df.empty else 0),
        ("Total Assets", round(bs.get("total_assets", 0), 2) if bs else 0),
        ("Total Liabilities", round(bs.get("total_liabilities", 0), 2) if bs else 0),
        ("Total Equity", round(bs.get("total_equity", 0), 2) if bs else 0),
        ("Balance Difference", round(bs.get("difference", 0), 2) if bs else 0),
        ("Red Flag Errors", sum(1 for f in (flags or []) if f.get("type") == "error")),
        ("Red Flag Warnings", sum(1 for f in (flags or []) if f.get("type") == "warning")),
    ]
    for i, (k, v) in enumerate(summary_data, 1):
        ws_sum.cell(row=i, column=1, value=k).font = Font(bold=True, name="Arial")
        ws_sum.cell(row=i, column=2, value=v)
    _auto_width(ws_sum)

    # ── Metadata ──────────────────────────────────────────────────
    ws_meta = wb.create_sheet("Metadata")
    meta = [
        ("App", "Financial Analyzer v1.0"),
        ("Generated", now),
        ("Engine", "Python / pandas / openpyxl"),
        ("Mapping Source", "Built-in (sample_mapping.json)"),
    ]
    for i, (k, v) in enumerate(meta, 1):
        ws_meta.cell(row=i, column=1, value=k).font = Font(bold=True, name="Arial")
        ws_meta.cell(row=i, column=2, value=v)
    _auto_width(ws_meta)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─────────────────────────── JSON export ─────────────────────────

def build_json_export(
    raw_df: pd.DataFrame,
    mapp_df: pd.DataFrame,
    bs: dict,
    flags: list[dict],
) -> str:
    def safe_df(df):
        if df is None or df.empty:
            return []
        return json.loads(df.to_json(orient="records", force_ascii=False))

    def safe_val(v):
        if v is None:
            return None
        if isinstance(v, float) and np.isnan(v):
            return None
        if isinstance(v, (np.integer,)):
            return int(v)
        if isinstance(v, (np.floating,)):
            return float(v)
        return v

    bs_serial = {}
    if bs:
        for k, v in bs.items():
            if isinstance(v, pd.DataFrame):
                bs_serial[k] = safe_df(v)
            else:
                bs_serial[k] = safe_val(v)

    payload = {
        "generated_at": datetime.now().isoformat(),
        "summary": {
            "total_accounts": int(len(raw_df)) if raw_df is not None else 0,
            "total_assets": safe_val(bs.get("total_assets")) if bs else None,
            "total_liabilities": safe_val(bs.get("total_liabilities")) if bs else None,
            "total_equity": safe_val(bs.get("total_equity")) if bs else None,
            "difference": safe_val(bs.get("difference")) if bs else None,
        },
        "red_flags": flags or [],
        "balance_sheet": bs_serial,
        "mapp": safe_df(mapp_df),
    }
    return json.dumps(payload, ensure_ascii=False, indent=2, default=str)
